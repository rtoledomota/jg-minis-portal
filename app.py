# Fix SQLite para Cloudflare Pages
try:
    import pysqlite3 as sqlite3
except ImportError:
    import sqlite3

import json
from flask import Flask, request, redirect, url_for, session, render_template_string, flash, send_file, Response
from functools import wraps
import os
import bcrypt
import re
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sua_chave_secreta_aqui')

# 📁 Persistência total - Sobrevive reinícios do Cloudflare
PERSIST_DIR = Path("/tmp") / "JG_MINIS_PERSIST_v4"
PERSIST_DIR.mkdir(parents=True, exist_ok=True)

DB_FILE = PERSIST_DIR / "database.db"
BACKUP_FILE = PERSIST_DIR / "backup_v4.json"

WHATSAPP_NUMERO = os.environ.get('WHATSAPP_NUMERO', '5511999999999')
GOOGLE_SHEETS_ID = os.environ.get('GOOGLE_SHEETS_ID', '1sxlvo6j-UTB0xXuyivzWnhRuYvpJFcH2smL4ZzHTUps')
BACKUP_SHEETS_ID = os.environ.get('BACKUP_SHEETS_ID', '1avMoEA0WddQ7dW92X2NORo-cJSwJb7cpinjsuZMMZqI')  # Sua planilha de backup
GOOGLE_SHEETS_SHEET = 'Miniaturas'
BACKUP_RESERVAS_SHEET = 'Backups_Reservas'
BACKUP_USUARIOS_SHEET = 'Usuarios_Backup'
LOGO_URL = "https://i.imgur.com/Yp1OiWB.jpeg"

# Configurações de Email (fallback simulação)
EMAIL_SENDER = os.environ.get('EMAIL_SENDER', 'seu_email@gmail.com')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', 'sua_senha_app')
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))

print(f"🚀 JG MINIS v4.2 - Persistência + Ordenação + Exports + Backups Planilhas: {DB_FILE}")

def get_db_connection():
    """Conexão com banco persistente"""
    return sqlite3.connect(str(DB_FILE))

def create_json_backup():
    """Backup automático JSON - Salva reservas e alterações"""
    try:
        conn = get_db_connection()
        c = conn.cursor()
        
        backup_data = {
            'timestamp': datetime.now().isoformat(),
            'users': c.execute("SELECT * FROM users").fetchall(),
            'miniaturas': c.execute("SELECT * FROM miniaturas").fetchall(),
            'reservations': c.execute("SELECT * FROM reservations").fetchall(),
            'waitlist': c.execute("SELECT * FROM waitlist").fetchall()
        }
        
        with open(str(BACKUP_FILE), 'w') as f:
            json.dump(backup_data, f, default=str)
        
        conn.close()
        print(f"💾 JSON Backup criado: {len(backup_data['reservations'])} reservas")
        return True
    except Exception as e:
        print(f"⚠️ Erro no JSON backup: {e}")
        return False

def restore_backup():
    """Restauração automática - Recupera reservas perdidas"""
    if BACKUP_FILE.exists():
        try:
            with open(str(BACKUP_FILE), 'r') as f:
                backup_data = json.load(f)
            
            conn = get_db_connection()
            c = conn.cursor()
            
            # Limpa e restaura (preserva admin)
            c.execute("DELETE FROM waitlist")
            c.execute("DELETE FROM reservations")
            c.execute("DELETE FROM miniaturas")
            c.execute("DELETE FROM users WHERE email != 'admin@jgminis.com.br'")
            
            # Restaura
            for user in backup_data.get('users', []):
                if user[2] != 'admin@jgminis.com.br':
                    c.execute("INSERT OR REPLACE INTO users VALUES (?, ?, ?, ?, ?, ?)", user)
            
            for mini in backup_data.get('miniaturas', []):
                # Compatibilidade com created_at
                if len(mini) == 8: # Versão antiga sem created_at
                    mini = (*mini, datetime.now().isoformat())
                c.execute("INSERT OR REPLACE INTO miniaturas VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", mini)
            
            for res in backup_data.get('reservations', []):
                c.execute("INSERT OR REPLACE INTO reservations VALUES (?, ?, ?, ?, ?, ?, ?)", res)
            
            for wl in backup_data.get('waitlist', []):
                c.execute("INSERT OR REPLACE INTO waitlist VALUES (?, ?, ?, ?, ?, ?)", wl)
            
            conn.commit()
            conn.close()
            
            num_reservas = len(backup_data.get('reservations', []))
            print(f"✅ Restaurado: {num_reservas} reservas do backup")
            
            # BACKUP_FILE.unlink() # Não apagar para manter persistência
            return True
        except Exception as e:
            print(f"⚠️ Erro na restauração: {e}")
            return False
    return False

def export_to_excel_reservas():
    """Export reservas para Excel - Automático (buffer) ou manual (download)"""
    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('''SELECT u.name as usuario, m.name as miniatura, r.quantity, r.reservation_date, r.status, m.price, (r.quantity * m.price) as total
                     FROM reservations r 
                     JOIN users u ON r.user_id = u.id 
                     JOIN miniaturas m ON r.miniatura_id = m.id 
                     ORDER BY r.reservation_date DESC''')
        reservas = c.fetchall()
        conn.close()
        
        if not reservas:
            print("Nenhuma reserva para exportar")
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservas"
        
        headers = ['Usuário', 'Miniatura', 'Quantidade', 'Data Reserva', 'Status', 'Preço Unitário', 'Total']
        ws.append(headers)
        
        # Estilo cabeçalhos (azul, bold, branco)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        for col in range(1, 8):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for row_data in reservas:
            ws.append(row_data)
            # Formatação total
            ws.cell(ws.max_row, column=7).number_format = 'R$ #,##0.00'
            # Cor por status
            status_cell = ws.cell(ws.max_row, column=5)
            status = row_data[4]
            if status == 'confirmed':
                status_cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            elif status == 'cancelled':
                status_cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        
        # Auto-ajuste colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Excel backup reservas criado")
        return buffer
    except Exception as e:
        print(f"⚠️ Erro no Excel export reservas: {e}")
        return None

def export_to_excel_usuarios():
    """Export usuários para Excel"""
    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('SELECT name, email, phone, is_admin FROM users ORDER BY name')
        usuarios = c.fetchall()
        conn.close()
        
        if not usuarios:
            print("Nenhum usuário para exportar")
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuários"
        
        headers = ['Nome', 'Email', 'Telefone', 'Admin']
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for row_data in usuarios:
            ws.append([row_data[0], row_data[1], row_data[2], 'Sim' if row_data[3] else 'Não'])
        
        # Auto-ajuste
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Excel backup usuários criado")
        return buffer
    except Exception as e:
        print(f"⚠️ Erro no Excel export usuários: {e}")
        return None

def get_google_sheets():
    try:
        credentials_json = os.environ.get('GOOGLE_SHEETS_CREDENTIALS', '{}')
        credentials_dict = json.loads(credentials_json)
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
        gc = gspread.authorize(credentials)
        return gc
    except Exception as e:
        print(f"Erro ao conectar ao Google Sheets: {e}")
        return None

def backup_to_google():
    """Backup automático para sua planilha Google - Reservas e Usuários"""
    gc = get_google_sheets()
    if not gc:
        print("⚠️ Sem credenciais Google - Backup Sheets pulado")
        return False
    
    try:
        sheet = gc.open_by_key(BACKUP_SHEETS_ID)
        
        # Aba reservas (cria se não existir)
        try:
            ws_reservas = sheet.worksheet(BACKUP_RESERVAS_SHEET)
        except gspread.WorksheetNotFound:
            ws_reservas = sheet.add_worksheet(title=BACKUP_RESERVAS_SHEET, rows=1000, cols=7)
        
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('''SELECT u.name, m.name, r.quantity, r.reservation_date, r.status, m.price, (r.quantity * m.price)
                     FROM reservations r JOIN users u ON r.user_id = u.id JOIN miniaturas m ON r.miniatura_id = m.id
                     ORDER BY r.reservation_date DESC''')
        reservas = c.fetchall()
        conn.close()
        
        if reservas:
            ws_reservas.clear()
            ws_reservas.append_row(['Usuário', 'Miniatura', 'Quantidade', 'Data Reserva', 'Status', 'Preço Unitário', 'Total'])
            ws_reservas.append_rows(reservas)
            # Formatação cabeçalho
            ws_reservas.format('A1:G1', {"backgroundColor": {"red": 0.23, "green": 0.51, "blue": 0.96}, "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}}})
            print(f"✅ Backup Google Reservas: {len(reservas)} linhas atualizadas na planilha")
        
        # Aba usuários
        try:
            ws_usuarios = sheet.worksheet(BACKUP_USUARIOS_SHEET)
        except gspread.WorksheetNotFound:
            ws_usuarios = sheet.add_worksheet(title=BACKUP_USUARIOS_SHEET, rows=1000, cols=4)
        
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('SELECT name, email, phone, is_admin FROM users ORDER BY name')
        usuarios = c.fetchall()
        conn.close()
        
        if usuarios:
            ws_usuarios.clear()
            ws_usuarios.append_row(['Nome', 'Email', 'Telefone', 'Admin'])
            ws_usuarios.append_rows([[u[0], u[1], u[2], 'Sim' if u[3] else 'Não'] for u in usuarios])
            ws_usuarios.format('A1:D1', {"backgroundColor": {"red": 0.54, "green": 0.36, "blue": 0.96}, "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}}})
            print(f"✅ Backup Google Usuários: {len(usuarios)} linhas atualizadas")
        
        return True
    except Exception as e:
        print(f"⚠️ Erro no backup Google: {e}")
        return False

def sync_to_sheets():
    """Sincroniza alterações de miniaturas com a planilha principal (atualiza estoque, novas inserções)"""
    gc = get_google_sheets()
    if not gc:
        print("⚠️ Sem credenciais Google - Sincronização Sheets pulada")
        return False
    
    try:
        sheet = gc.open_by_key(GOOGLE_SHEETS_ID)
        ws = sheet.worksheet(GOOGLE_SHEETS_SHEET)
        
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('SELECT image_url, name, arrival_date, stock, price, observations, max_reservations_per_user FROM miniaturas')
        miniaturas = c.fetchall()
        conn.close()
        
        if miniaturas:
            ws.clear()
            ws.append_row(['Image URL', 'Nome', '', 'Arrival Date', 'Stock', 'Price', 'Observations', 'Max Reservations'])
            ws.append_rows([[m[0], m[1], '', m[2], m[3], m[4], m[5], m[6]] for m in miniaturas])
            print(f"✅ Sincronização Sheets: {len(miniaturas)} miniaturas atualizadas na planilha principal")
        
        return True
    except Exception as e:
        print(f"⚠️ Erro na sincronização Sheets: {e}")
        return False

def enviar_email(destinatario, assunto, corpo_html):
    try:
        if EMAIL_SENDER == 'seu_email@gmail.com' or not EMAIL_PASSWORD:
            print(f"📧 SIMULAÇÃO: Email para {destinatario} - {assunto}")
            return True
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = assunto
        msg['From'] = EMAIL_SENDER
        msg['To'] = destinatario
        
        parte_html = MIMEText(corpo_html, 'html', 'utf-8')
        msg.attach(parte_html)
        
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
        
        print(f"✅ Email enviado para {destinatario}")
        return True
    except Exception as e:
        print(f"⚠️ Erro ao enviar email: {e}")
        return False

def init_db():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, email TEXT UNIQUE NOT NULL, phone TEXT, password TEXT NOT NULL, is_admin INTEGER DEFAULT 0)''')
    c.execute('''CREATE TABLE IF NOT EXISTS miniaturas (id INTEGER PRIMARY KEY AUTOINCREMENT, image_url TEXT NOT NULL, name TEXT NOT NULL, arrival_date TEXT, stock INTEGER DEFAULT 0, price REAL DEFAULT 0.0, observations TEXT, max_reservations_per_user INTEGER DEFAULT 1, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS reservations (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, miniatura_id INTEGER NOT NULL, quantity INTEGER NOT NULL, reservation_date TEXT NOT NULL, status TEXT DEFAULT 'confirmed', confirmacao_enviada INTEGER DEFAULT 0, FOREIGN KEY (user_id) REFERENCES users(id), FOREIGN KEY (miniatura_id) REFERENCES miniaturas(id))''')
    c.execute('''CREATE TABLE IF NOT EXISTS waitlist (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, miniatura_id INTEGER NOT NULL, email TEXT NOT NULL, notification_sent INTEGER DEFAULT 0, request_date TEXT NOT NULL, FOREIGN KEY (user_id) REFERENCES users(id), FOREIGN KEY (miniatura_id) REFERENCES miniaturas(id))''')
    conn.commit()
    conn.close()
    print("✅ BD inicializado com persistência e created_at para ordenação")

def load_initial_data():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE email = 'admin@jgminis.com.br'")
    if not c.fetchone():
        hashed_password = bcrypt.hashpw('admin123'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        c.execute("INSERT INTO users (name, email, phone, password, is_admin) VALUES (?, ?, ?, ?, ?)", ('Admin', 'admin@jgminis.com.br', '5511999999999', hashed_password, 1))
        print("✅ Usuário admin adicionado.")
    c.execute("SELECT * FROM users WHERE email = 'usuario@example.com'")
    if not c.fetchone():
        hashed_password = bcrypt.hashpw('usuario123'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        c.execute("INSERT INTO users (name, email, phone, password, is_admin) VALUES (?, ?, ?, ?, ?)", ('Usuário Teste', 'usuario@example.com', '5511988888888', hashed_password, 0))
        print("✅ Usuário teste adicionado.")
    conn.commit()
    conn.close()

def load_miniaturas_from_sheets():
    gc = get_google_sheets()
    if not gc:
        print("Não foi possível conectar ao Google Sheets")
        return
    try:
        sheet = gc.open_by_key(GOOGLE_SHEETS_ID).worksheet(GOOGLE_SHEETS_SHEET)
        rows = sheet.get_all_values()
        if not rows:
            print("Planilha vazia")
            return
        conn = get_db_connection()
        c = conn.cursor()
        # c.execute("DELETE FROM miniaturas") # Não deletar, apenas atualizar ou inserir
        for row in rows[1:]:
            if len(row) >= 8:
                try:
                    image_url = row[0]
                    name = row[1]
                    arrival_date = row[3]
                    stock = int(row[4]) if row[4] else 0
                    price = float(row[5]) if row[5] else 0.0
                    observations = row[6]
                    max_reservations = int(row[7]) if row[7] else 1
                    created_at = datetime.now().isoformat()  # Adiciona created_at
                    
                    # Verifica se a miniatura já existe pelo nome
                    c.execute("SELECT id FROM miniaturas WHERE name = ?", (name,))
                    existing_mini = c.fetchone()
                    
                    if existing_mini:
                        c.execute("UPDATE miniaturas SET image_url=?, arrival_date=?, stock=?, price=?, observations=?, max_reservations_per_user=? WHERE id=?", 
                                  (image_url, arrival_date, stock, price, observations, max_reservations, existing_mini[0]))
                    else:
                        c.execute("INSERT INTO miniaturas (image_url, name, arrival_date, stock, price, observations, max_reservations_per_user, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (image_url, name, arrival_date, stock, price, observations, max_reservations, created_at))
                except Exception as e:
                    print(f"Erro ao inserir/atualizar linha da planilha: {e}")
                    continue
        conn.commit()
        conn.close()
        print(f"✅ {len(rows)-1} miniaturas carregadas/atualizadas do Google Sheets")
        # create_json_backup() # Não chamar aqui para evitar loop/timeout
    except Exception as e:
        print(f"Erro ao carregar planilha: {e}")

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'error')
            return redirect(url_for('login'))
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('SELECT id, name, email, phone, is_admin FROM users WHERE id = ?', (session['user_id'],))
        user = c.fetchone()
        conn.close()
        if user:
            request.user = {'user_id': user[0], 'name': user[1], 'email': user[2], 'phone': user[3], 'is_admin': bool(user[4])}
        else:
            session.pop('user_id', None)
            flash('Sua sessão expirou ou usuário não encontrado.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not request.user.get('is_admin'):
            flash('Acesso negado: Você não tem permissões de administrador.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        phone = request.form['phone']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            flash('Formato de e-mail inválido.', 'error')
            return redirect(url_for('register'))
        if password != confirm_password:
            flash('As senhas não coincidem.', 'error')
            return redirect(url_for('register'))
        if len(password) < 6:
            flash('A senha deve ter pelo menos 6 caracteres.', 'error')
            return redirect(url_for('register'))
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute("INSERT INTO users (name, email, phone, password) VALUES (?, ?, ?, ?)", (name, email, phone, hashed_password))
            conn.commit()
            create_json_backup()  # Backup após novo usuário
            flash('Registro bem-sucedido! Faça login.', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('E-mail já registrado.', 'error')
        finally:
            conn.close()
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Registrar - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen flex items-center justify-center">
        <div class="bg-gradient-to-b from-slate-800 to-black rounded-xl border-2 border-red-600 shadow-2xl p-8 max-w-md w-full">
            <div class="flex justify-center mb-6">
                <img src="''' + LOGO_URL + '''" class="h-16 rounded-full border-2 border-blue-400" alt="JG MINIS">
            </div>
            <h2 class="text-3xl font-black text-blue-400 mb-6 text-center">Registrar</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            <form method="POST" action="/register" class="space-y-4">
                <div>
                    <label for="name" class="block text-slate-300 font-bold mb-1">Nome:</label>
                    <input type="text" id="name" name="name" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                </div>
                <div>
                    <label for="email" class="block text-slate-300 font-bold mb-1">E-mail:</label>
                    <input type="email" id="email" name="email" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                </div>
                <div>
                    <label for="phone" class="block text-slate-300 font-bold mb-1">Telefone:</label>
                    <input type="text" id="phone" name="phone" class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                </div>
                <div>
                    <label for="password" class="block text-slate-300 font-bold mb-1">Senha:</label>
                    <input type="password" id="password" name="password" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                </div>
                <div>
                    <label for="confirm_password" class="block text-slate-300 font-bold mb-1">Confirmar Senha:</label>
                    <input type="password" id="confirm_password" name="confirm_password" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                </div>
                <button type="submit" class="w-full bg-gradient-to-r from-blue-600 to-red-600 text-white font-bold py-2 rounded-lg hover:from-blue-700 hover:to-red-700 transition duration-300">Registrar</button>
            </form>
            <p class="text-center text-slate-400 mt-6">Já tem uma conta? <a href="/login" class="text-blue-400 hover:underline">Faça Login</a></p>
        </div>
    </body>
    </html>
    ''', LOGO_URL=LOGO_URL)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('SELECT id, password FROM users WHERE email = ?', (email,))
        user_data = c.fetchone()
        conn.close()
        if user_data and bcrypt.checkpw(password.encode('utf-8'), user_data[1].encode('utf-8')):
            session['user_id'] = user_data[0]
            flash('Login bem-sucedido!', 'success')
            return redirect(url_for('index'))
        else:
            flash('E-mail ou senha inválidos.', 'error')
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Login - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen flex items-center justify-center">
        <div class="bg-gradient-to-b from-slate-800 to-black rounded-xl border-2 border-red-600 shadow-2xl p-8 max-w-md w-full">
            <div class="flex justify-center mb-6">
                <img src="''' + LOGO_URL + '''" class="h-16 rounded-full border-2 border-blue-400" alt="JG MINIS">
            </div>
            <h2 class="text-3xl font-black text-blue-400 mb-6 text-center">Login</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            <form method="POST" action="/login" class="space-y-4">
                <div>
                    <label for="email" class="block text-slate-300 font-bold mb-1">E-mail:</label>
                    <input type="email" id="email" name="email" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                </div>
                <div>
                    <label for="password" class="block text-slate-300 font-bold mb-1">Senha:</label>
                    <input type="password" id="password" name="password" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                </div>
                <button type="submit" class="w-full bg-gradient-to-r from-blue-600 to-red-600 text-white font-bold py-2 rounded-lg hover:from-blue-700 hover:to-red-700 transition duration-300">Login</button>
            </form>
            <p class="text-center text-slate-400 mt-6">Não tem conta? <a href="/register" class="text-blue-400 hover:underline">Registre-se</a></p>
        </div>
    </body>
    </html>
    ''', LOGO_URL=LOGO_URL)

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    flash('Logout realizado com sucesso!', 'success')
    return redirect(url_for('login'))

@app.route('/', defaults={'sort': 'name', 'order': 'asc'})
@app.route('/<sort>/<order>')
@login_required
def index(sort, order):
    conn = get_db_connection()
    c = conn.cursor()
    
    # Ordenação dinâmica
    order_by_map = {
        'name': 'name',
        'arrival_date': 'arrival_date',
        'created_at': 'created_at',
        'stock': 'stock'
    }
    safe_sort = order_by_map.get(sort, 'name')
    safe_order = 'ASC' if order == 'asc' else 'DESC'

    c.execute(f"SELECT * FROM miniaturas ORDER BY {safe_sort} {safe_order}")
    miniaturas = c.fetchall()
    conn.close()
    
    # Opções de ordenação para o template
    sort_options = [
        ('name', 'Nome'), 
        ('arrival_date', 'Previsão de Chegada'), 
        ('created_at', 'Data de Inclusão'), 
        ('stock', 'Disponibilidade')
    ]
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Catálogo - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/minhas-reservas" class="text-slate-300 hover:text-blue-400">Minhas Reservas</a>
                    <a href="/perfil" class="text-slate-300 hover:text-blue-400">Perfil</a>
                    {% if request.user.is_admin %}
                        <a href="/admin" class="text-slate-300 hover:text-blue-400">Admin</a>
                    {% endif %}
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Catálogo de Miniaturas</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            
            <!-- Ordenação -->
            <div class="bg-slate-800 rounded-lg p-4 mb-6 flex flex-wrap items-center gap-4">
                <form method="GET" action="/" class="flex items-center gap-4">
                    <label for="sort" class="text-slate-300 font-bold">Ordenar por:</label>
                    <select id="sort" name="sort" class="bg-slate-700 text-white px-4 py-2 rounded-lg border border-blue-600 focus:outline-none focus:border-red-500">
                        {% for s, label in sort_options %}
                            <option value="{{ s }}" {% if s == sort %}selected{% endif %}>{{ label }}</option>
                        {% endfor %}
                    </select>
                    <select id="order" name="order" class="bg-slate-700 text-white px-4 py-2 rounded-lg border border-blue-600 focus:outline-none focus:border-red-500">
                        <option value="asc" {% if order == 'asc' %}selected{% endif %}>Crescente</option>
                        <option value="desc" {% if order == 'desc' %}selected{% endif %}>Decrescente</option>
                    </select>
                    <button type="submit" class="bg-blue-600 text-white px-4 py-2 rounded-lg hover:bg-blue-700 transition duration-300">Aplicar</button>
                </form>
            </div>
            
            <div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-6">
                {% for mini in miniaturas %}
                    <div class="bg-slate-800 rounded-lg p-4 border-2 border-blue-600 flex flex-col">
                        <img src="{{ mini[1] }}" class="w-full h-48 object-cover rounded-lg mb-4" alt="{{ mini[2] }}">
                        <h3 class="text-xl font-bold text-white mb-2">{{ mini[2] }}</h3>
                        <p class="text-slate-300 mb-2">Previsão: {{ mini[3] }}</p>
                        <p class="text-slate-300 mb-2">Estoque: {{ mini[4] }}</p>
                        <p class="text-green-400 font-bold mb-4">R$ {{ "%.2f"|format(mini[5]) }}</p>
                        <p class="text-slate-300 mb-4 flex-grow">{{ mini[6] }}</p>
                        {% if mini[4] > 0 %}
                            <a href="/reservar/{{ mini[0] }}" class="bg-green-600 text-white px-4 py-2 rounded-lg hover:bg-green-700 block text-center mt-auto">Reservar</a>
                        {% else %}
                            <a href="https://wa.me/{{ WHATSAPP_NUMERO }}?text=Olá, tenho interesse na miniatura {{ mini[2] }} e gostaria de ser notificado quando houver estoque." target="_blank" class="bg-blue-600 text-white px-4 py-2 rounded-lg hover:bg-blue-700 block text-center mt-auto">Notificar Estoque</a>
                        {% endif %}
                    </div>
                {% endfor %}
            </div>
        </div>
    </body>
    </html>
    ''', miniaturas=miniaturas, sort=sort, order=order, sort_options=sort_options, LOGO_URL=LOGO_URL, WHATSAPP_NUMERO=WHATSAPP_NUMERO)

@app.route('/reservar/<int:mini_id>', methods=['GET', 'POST'])
@login_required
def reservar(mini_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT id, name, stock, price, max_reservations_per_user FROM miniaturas WHERE id = ?", (mini_id,))
    mini = c.fetchone()
    
    if not mini:
        conn.close()
        flash('Miniatura não encontrada.', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        try:
            quantity = int(request.form['quantity'])
            if quantity <= 0:
                flash('Quantidade deve ser positiva.', 'error')
                return redirect(url_for('reservar', mini_id=mini_id))
        except ValueError:
            flash('Quantidade inválida.', 'error')
            return redirect(url_for('reservar', mini_id=mini_id))

        if quantity > mini[2]: # mini[2] é o stock
            flash(f'Estoque insuficiente. Disponível: {mini[2]}', 'error')
        else:
            c.execute("SELECT SUM(quantity) FROM reservations WHERE user_id = ? AND miniatura_id = ? AND status = 'confirmed'", (session['user_id'], mini_id))
            existing_reservations = c.fetchone()[0] or 0
            
            if existing_reservations + quantity > mini[4]: # mini[4] é max_reservations_per_user
                flash(f'Limite de reservas excedido para esta miniatura. Você já reservou {existing_reservations}. Máximo permitido: {mini[4]}', 'error')
            else:
                reservation_date = datetime.now().isoformat()
                c.execute("INSERT INTO reservations (user_id, miniatura_id, quantity, reservation_date) VALUES (?, ?, ?, ?)", (session['user_id'], mini_id, quantity, reservation_date))
                c.execute("UPDATE miniaturas SET stock = stock - ? WHERE id = ?", (quantity, mini_id))
                conn.commit()
                
                create_json_backup() # Backup local
                backup_to_google() # Backup Google Sheets
                sync_to_sheets() # Sincroniza planilha principal
                
                # Email de confirmação
                corpo_html = f'''
                <h1>Reserva Confirmada - JG MINIS</h1>
                <p>Olá {request.user['name']},</p>
                <p>Sua reserva para a miniatura <strong>{mini[1]}</strong> foi confirmada!</p>
                <p><strong>Detalhes da Reserva:</strong></p>
                <ul>
                    <li>Miniatura: {mini[1]}</li>
                    <li>Quantidade: {quantity}</li>
                    <li>Preço Unitário: R$ {mini[3]:.2f}</li>
                    <li>Total: R$ {(quantity * mini[3]):.2f}</li>
                    <li>Data da Reserva: {datetime.now().strftime('%d/%m/%Y %H:%M')}</li>
                </ul>
                <p>Agradecemos a sua preferência!</p>
                <p>Atenciosamente,<br>Equipe JG MINIS</p>
                '''
                enviar_email(request.user['email'], 'Confirmação de Reserva JG MINIS', corpo_html)
                flash('Reserva realizada com sucesso e e-mail de confirmação enviado!', 'success')
                return redirect(url_for('minhas_reservas'))
    conn.close()
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Reservar - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/" class="text-slate-300 hover:text-blue-400">Catálogo</a>
                    <a href="/minhas-reservas" class="text-slate-300 hover:text-blue-400">Minhas Reservas</a>
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Reservar {{ mini[1] }}</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            <div class="bg-slate-800 rounded-lg p-6 max-w-md mx-auto border-2 border-blue-600">
                <img src="{{ mini_image_url }}" class="w-full h-48 object-cover rounded-lg mb-4" alt="{{ mini[1] }}">
                <p class="text-slate-300 mb-2">Estoque disponível: {{ mini[2] }}</p>
                <p class="text-slate-300 mb-2">Preço unitário: R$ {{ "%.2f"|format(mini[3]) }}</p>
                <p class="text-slate-300 mb-4">Máximo por usuário: {{ mini[4] }}</p>
                <form method="POST" class="space-y-4">
                    <div>
                        <label for="quantity" class="block text-white font-bold mb-2">Quantidade:</label>
                        <input type="number" id="quantity" name="quantity" min="1" max="{{ mini[2] }}" value="1" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <button type="submit" class="w-full bg-green-600 text-white font-bold py-2 rounded-lg hover:bg-green-700 transition duration-300">Confirmar Reserva</button>
                </form>
            </div>
        </div>
    </body>
    </html>
    ''', mini=mini, mini_image_url=mini[0], LOGO_URL=LOGO_URL) # mini[0] é image_url

@app.route('/minhas-reservas')
@login_required
def minhas_reservas():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''SELECT r.id, m.name, r.quantity, m.price, r.reservation_date, r.status, m.image_url
                 FROM reservations r
                 JOIN miniaturas m ON r.miniatura_id = m.id
                 WHERE r.user_id = ?
                 ORDER BY r.reservation_date DESC''', (session['user_id'],))
    reservas = c.fetchall()
    conn.close()
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Minhas Reservas - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/" class="text-slate-300 hover:text-blue-400">Catálogo</a>
                    <a href="/perfil" class="text-slate-300 hover:text-blue-400">Perfil</a>
                    {% if request.user.is_admin %}
                        <a href="/admin" class="text-slate-300 hover:text-blue-400">Admin</a>
                    {% endif %}
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Minhas Reservas</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            
            {% if reservas %}
                <div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-6">
                    {% for res in reservas %}
                        <div class="bg-slate-800 rounded-lg p-4 border-2 border-blue-600 flex flex-col">
                            <img src="{{ res[6] }}" class="w-full h-48 object-cover rounded-lg mb-4" alt="{{ res[1] }}">
                            <h3 class="text-xl font-bold text-white mb-2">{{ res[1] }}</h3>
                            <p class="text-slate-300 mb-2">Quantidade: {{ res[2] }}</p>
                            <p class="text-slate-300 mb-2">Preço Unitário: R$ {{ "%.2f"|format(res[3]) }}</p>
                            <p class="text-slate-300 mb-2">Total: R$ {{ "%.2f"|format(res[2] * res[3]) }}</p>
                            <p class="text-slate-300 mb-2">Data da Reserva: {{ res[4] }}</p>
                            <p class="text-{{ 'green' if res[5] == 'confirmed' else 'red' }}-400 font-bold mb-4">Status: {{ res[5] }}</p>
                            {% if res[5] == 'confirmed' %}
                                <a href="/cancelar-reserva/{{ res[0] }}" class="bg-red-600 text-white px-4 py-2 rounded-lg hover:bg-red-700 block text-center mt-auto">Cancelar Reserva</a>
                            {% endif %}
                        </div>
                    {% endfor %}
                </div>
            {% else %}
                <p class="text-white text-center text-lg">Você ainda não fez nenhuma reserva.</p>
                <p class="text-white text-center text-lg mt-2"><a href="/" class="text-blue-400 hover:underline">Explore o catálogo!</a></p>
            {% endif %}
        </div>
    </body>
    </html>
    ''', reservas=reservas, LOGO_URL=LOGO_URL)

@app.route('/cancelar-reserva/<int:res_id>')
@login_required
def cancelar_reserva(res_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT user_id, miniatura_id, quantity, status FROM reservations WHERE id = ?", (res_id,))
    reserva = c.fetchone()
    
    if not reserva:
        conn.close()
        flash('Reserva não encontrada.', 'error')
        return redirect(url_for('minhas_reservas'))
    
    if reserva[0] != session['user_id'] and not request.user['is_admin']:
        conn.close()
        flash('Você não tem permissão para cancelar esta reserva.', 'error')
        return redirect(url_for('minhas_reservas'))
    
    if reserva[3] == 'cancelled':
        conn.close()
        flash('Esta reserva já foi cancelada.', 'error')
        return redirect(url_for('minhas_reservas'))

    c.execute("UPDATE reservations SET status = 'cancelled' WHERE id = ?", (res_id,))
    c.execute("UPDATE miniaturas SET stock = stock + ? WHERE id = ?", (reserva[2], reserva[1]))
    conn.commit()
    
    create_json_backup() # Backup local
    backup_to_google() # Backup Google Sheets
    sync_to_sheets() # Sincroniza planilha principal
    
    flash('Reserva cancelada com sucesso!', 'success')
    conn.close()
    return redirect(url_for('minhas_reservas'))

@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT name, email, phone FROM users WHERE id = ?', (session['user_id'],))
    user_data = c.fetchone()
    conn.close()

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        phone = request.form['phone']
        
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            flash('Formato de e-mail inválido.', 'error')
            return redirect(url_for('perfil'))

        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute("UPDATE users SET name = ?, email = ?, phone = ? WHERE id = ?", (name, email, phone, session['user_id']))
            conn.commit()
            create_json_backup() # Backup local
            flash('Perfil atualizado com sucesso!', 'success')
            return redirect(url_for('perfil'))
        except sqlite3.IntegrityError:
            flash('Este e-mail já está em uso.', 'error')
        finally:
            conn.close()
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Meu Perfil - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/" class="text-slate-300 hover:text-blue-400">Catálogo</a>
                    <a href="/minhas-reservas" class="text-slate-300 hover:text-blue-400">Minhas Reservas</a>
                    {% if request.user.is_admin %}
                        <a href="/admin" class="text-slate-300 hover:text-blue-400">Admin</a>
                    {% endif %}
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Meu Perfil</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            <div class="bg-slate-800 rounded-lg p-6 max-w-md mx-auto border-2 border-blue-600">
                <form method="POST" class="space-y-4">
                    <div>
                        <label for="name" class="block text-slate-300 font-bold mb-1">Nome:</label>
                        <input type="text" id="name" name="name" value="{{ user_data[0] }}" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="email" class="block text-slate-300 font-bold mb-1">E-mail:</label>
                        <input type="email" id="email" name="email" value="{{ user_data[1] }}" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="phone" class="block text-slate-300 font-bold mb-1">Telefone:</label>
                        <input type="text" id="phone" name="phone" value="{{ user_data[2] }}" class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <button type="submit" class="w-full bg-gradient-to-r from-blue-600 to-red-600 text-white font-bold py-2 rounded-lg hover:from-blue-700 hover:to-red-700 transition duration-300">Atualizar Perfil</button>
                </form>
                <p class="text-center mt-4"><a href="/mudar-senha" class="text-blue-400 hover:underline">Mudar Senha</a></p>
            </div>
        </div>
    </body>
    </html>
    ''', user_data=user_data, LOGO_URL=LOGO_URL)

@app.route('/mudar-senha', methods=['GET', 'POST'])
@login_required
def mudar_senha():
    if request.method == 'POST':
        old_password = request.form['old_password']
        new_password = request.form['new_password']
        confirm_new_password = request.form['confirm_new_password']

        conn = get_db_connection()
        c = conn.cursor()
        c.execute('SELECT password FROM users WHERE id = ?', (session['user_id'],))
        user_password_hash = c.fetchone()[0]
        conn.close()

        if not bcrypt.checkpw(old_password.encode('utf-8'), user_password_hash.encode('utf-8')):
            flash('Senha antiga incorreta.', 'error')
        elif new_password != confirm_new_password:
            flash('As novas senhas não coincidem.', 'error')
        elif len(new_password) < 6:
            flash('A nova senha deve ter pelo menos 6 caracteres.', 'error')
        else:
            hashed_new_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            conn = get_db_connection()
            c = conn.cursor()
            c.execute("UPDATE users SET password = ? WHERE id = ?", (hashed_new_password, session['user_id']))
            conn.commit()
            create_json_backup() # Backup local
            conn.close()
            flash('Senha alterada com sucesso!', 'success')
            return redirect(url_for('perfil'))
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Mudar Senha - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/" class="text-slate-300 hover:text-blue-400">Catálogo</a>
                    <a href="/minhas-reservas" class="text-slate-300 hover:text-blue-400">Minhas Reservas</a>
                    <a href="/perfil" class="text-slate-300 hover:text-blue-400">Perfil</a>
                    {% if request.user.is_admin %}
                        <a href="/admin" class="text-slate-300 hover:text-blue-400">Admin</a>
                    {% endif %}
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Mudar Senha</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            <div class="bg-slate-800 rounded-lg p-6 max-w-md mx-auto border-2 border-blue-600">
                <form method="POST" class="space-y-4">
                    <div>
                        <label for="old_password" class="block text-slate-300 font-bold mb-1">Senha Antiga:</label>
                        <input type="password" id="old_password" name="old_password" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="new_password" class="block text-slate-300 font-bold mb-1">Nova Senha:</label>
                        <input type="password" id="new_password" name="new_password" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="confirm_new_password" class="block text-slate-300 font-bold mb-1">Confirmar Nova Senha:</label>
                        <input type="password" id="confirm_new_password" name="confirm_new_password" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <button type="submit" class="w-full bg-gradient-to-r from-blue-600 to-red-600 text-white font-bold py-2 rounded-lg hover:from-blue-700 hover:to-red-700 transition duration-300">Mudar Senha</button>
                </form>
            </div>
        </div>
    </body>
    </html>
    ''', LOGO_URL=LOGO_URL)

@app.route('/admin')
@login_required
@admin_required
def admin():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT * FROM miniaturas ORDER BY created_at DESC')
    miniaturas = c.fetchall()
    c.execute('''SELECT r.id, u.name, m.name, r.quantity, r.reservation_date, r.status
                 FROM reservations r JOIN users u ON r.user_id = u.id JOIN miniaturas m ON r.miniatura_id = m.id
                 ORDER BY r.reservation_date DESC''')
    reservations = c.fetchall()
    c.execute('''SELECT w.id, u.name, m.name, w.email, w.request_date
                 FROM waitlist w JOIN users u ON w.user_id = u.id JOIN miniaturas m ON w.miniatura_id = m.id
                 ORDER BY w.request_date DESC''')
    waitlist = c.fetchall()
    conn.close()
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Admin - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS Admin</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/" class="text-slate-300 hover:text-blue-400">Catálogo</a>
                    <a href="/minhas-reservas" class="text-slate-300 hover:text-blue-400">Minhas Reservas</a>
                    <a href="/perfil" class="text-slate-300 hover:text-blue-400">Perfil</a>
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Painel Administrativo</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}

            <div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-6 mb-8">
                <a href="/admin/add-miniatura" class="bg-green-600 text-white p-4 rounded-lg text-center font-bold hover:bg-green-700 transition duration-300">Adicionar Miniatura</a>
                <a href="/admin/gerenciar-usuarios" class="bg-blue-600 text-white p-4 rounded-lg text-center font-bold hover:bg-blue-700 transition duration-300">Gerenciar Usuários</a>
                <a href="/admin/export-reservas" class="bg-purple-600 text-white p-4 rounded-lg text-center font-bold hover:bg-purple-700 transition duration-300">Exportar Reservas (Excel)</a>
                <a href="/admin/export-usuarios" class="bg-purple-600 text-white p-4 rounded-lg text-center font-bold hover:bg-purple-700 transition duration-300">Exportar Usuários (Excel)</a>
                <a href="/admin/load-sheets" class="bg-yellow-600 text-white p-4 rounded-lg text-center font-bold hover:bg-yellow-700 transition duration-300">Carregar Miniaturas do Sheets</a>
                <a href="/admin/manual-backup" class="bg-indigo-600 text-white p-4 rounded-lg text-center font-bold hover:bg-indigo-700 transition duration-300">Backup Manual (JSON + Sheets)</a>
            </div>

            <h3 class="text-2xl font-bold text-white mb-4">Miniaturas Cadastradas</h3>
            {% if miniaturas %}
                <div class="overflow-x-auto bg-slate-800 rounded-lg border-2 border-blue-600 mb-8">
                    <table class="min-w-full divide-y divide-slate-700">
                        <thead class="bg-slate-700">
                            <tr>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">ID</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Imagem</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Nome</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Previsão</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Estoque</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Preço</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Ações</th>
                            </tr>
                        </thead>
                        <tbody class="bg-slate-800 divide-y divide-slate-700">
                            {% for mini in miniaturas %}
                                <tr>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-white">{{ mini[0] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300"><img src="{{ mini[1] }}" class="h-10 w-10 object-cover rounded-full"></td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ mini[2] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ mini[3] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ mini[4] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">R$ {{ "%.2f"|format(mini[5]) }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                        <a href="/admin/edit-miniatura/{{ mini[0] }}" class="text-indigo-400 hover:text-indigo-600 mr-4">Editar</a>
                                        <a href="/admin/delete-miniatura/{{ mini[0] }}" class="text-red-400 hover:text-red-600">Excluir</a>
                                    </td>
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            {% else %}
                <p class="text-white text-center text-lg mb-8">Nenhuma miniatura cadastrada.</p>
            {% endif %}

            <h3 class="text-2xl font-bold text-white mb-4">Reservas Ativas</h3>
            {% if reservations %}
                <div class="overflow-x-auto bg-slate-800 rounded-lg border-2 border-red-600 mb-8">
                    <table class="min-w-full divide-y divide-slate-700">
                        <thead class="bg-slate-700">
                            <tr>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">ID</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Usuário</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Miniatura</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Qtd</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Data</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Status</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Ações</th>
                            </tr>
                        </thead>
                        <tbody class="bg-slate-800 divide-y divide-slate-700">
                            {% for res in reservations %}
                                <tr>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-white">{{ res[0] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ res[1] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ res[2] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ res[3] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ res[4] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-{{ 'green' if res[5] == 'confirmed' else 'red' }}-400">{{ res[5] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                        {% if res[5] == 'confirmed' %}
                                            <a href="/cancelar-reserva/{{ res[0] }}" class="text-red-400 hover:text-red-600">Cancelar</a>
                                        {% endif %}
                                    </td>
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            {% else %}
                <p class="text-white text-center text-lg mb-8">Nenhuma reserva ativa.</p>
            {% endif %}

            <h3 class="text-2xl font-bold text-white mb-4">Lista de Espera</h3>
            {% if waitlist %}
                <div class="overflow-x-auto bg-slate-800 rounded-lg border-2 border-blue-600 mb-8">
                    <table class="min-w-full divide-y divide-slate-700">
                        <thead class="bg-slate-700">
                            <tr>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">ID</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Usuário</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Miniatura</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Email</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Data Pedido</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Ações</th>
                            </tr>
                        </thead>
                        <tbody class="bg-slate-800 divide-y divide-slate-700">
                            {% for item in waitlist %}
                                <tr>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-white">{{ item[0] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ item[1] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ item[2] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ item[3] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ item[4] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                        <a href="/admin/clear-waitlist/{{ item[0] }}" class="text-red-400 hover:text-red-600">Remover</a>
                                    </td>
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            {% else %}
                <p class="text-white text-center text-lg mb-8">Ninguém na lista de espera.</p>
            {% endif %}
        </div>
    </body>
    </html>
    ''', miniaturas=miniaturas, reservations=reservations, waitlist=waitlist, LOGO_URL=LOGO_URL)

@app.route('/admin/add-miniatura', methods=['GET', 'POST'])
@login_required
@admin_required
def add_miniatura():
    if request.method == 'POST':
        image_url = request.form['image_url']
        name = request.form['name']
        arrival_date = request.form['arrival_date']
        stock = int(request.form['stock'])
        price = float(request.form['price'])
        observations = request.form['observations']
        max_reservations_per_user = int(request.form['max_reservations_per_user'])
        created_at = datetime.now().isoformat()

        conn = get_db_connection()
        c = conn.cursor()
        c.execute("INSERT INTO miniaturas (image_url, name, arrival_date, stock, price, observations, max_reservations_per_user, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                  (image_url, name, arrival_date, stock, price, observations, max_reservations_per_user, created_at))
        conn.commit()
        create_json_backup() # Backup local
        backup_to_google() # Backup Google Sheets
        sync_to_sheets() # Sincroniza planilha principal
        conn.close()
        flash('Miniatura adicionada com sucesso!', 'success')
        return redirect(url_for('admin'))
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Adicionar Miniatura - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS Admin</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/admin" class="text-slate-300 hover:text-blue-400">Voltar para Admin</a>
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Adicionar Nova Miniatura</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            <div class="bg-slate-800 rounded-lg p-6 max-w-lg mx-auto border-2 border-blue-600">
                <form method="POST" class="space-y-4">
                    <div>
                        <label for="image_url" class="block text-slate-300 font-bold mb-1">URL da Imagem:</label>
                        <input type="text" id="image_url" name="image_url" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="name" class="block text-slate-300 font-bold mb-1">Nome da Miniatura:</label>
                        <input type="text" id="name" name="name" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="arrival_date" class="block text-slate-300 font-bold mb-1">Previsão de Chegada:</label>
                        <input type="text" id="arrival_date" name="arrival_date" class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="stock" class="block text-slate-300 font-bold mb-1">Estoque:</label>
                        <input type="number" id="stock" name="stock" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="price" class="block text-slate-300 font-bold mb-1">Preço:</label>
                        <input type="number" step="0.01" id="price" name="price" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="observations" class="block text-slate-300 font-bold mb-1">Observações:</label>
                        <textarea id="observations" name="observations" rows="3" class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500"></textarea>
                    </div>
                    <div>
                        <label for="max_reservations_per_user" class="block text-slate-300 font-bold mb-1">Máx. Reservas por Usuário:</label>
                        <input type="number" id="max_reservations_per_user" name="max_reservations_per_user" value="1" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <button type="submit" class="w-full bg-green-600 text-white font-bold py-2 rounded-lg hover:bg-green-700 transition duration-300">Adicionar Miniatura</button>
                </form>
            </div>
        </div>
    </body>
    </html>
    ''', LOGO_URL=LOGO_URL)

@app.route('/admin/edit-miniatura/<int:mini_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_miniatura(mini_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT image_url, name, arrival_date, stock, price, observations, max_reservations_per_user FROM miniaturas WHERE id = ?", (mini_id,))
    miniatura = c.fetchone()

    if not miniatura:
        conn.close()
        flash('Miniatura não encontrada.', 'error')
        return redirect(url_for('admin'))

    if request.method == 'POST':
        image_url = request.form['image_url']
        name = request.form['name']
        arrival_date = request.form['arrival_date']
        stock = int(request.form['stock'])
        price = float(request.form['price'])
        observations = request.form['observations']
        max_reservations_per_user = int(request.form['max_reservations_per_user'])

        c.execute("UPDATE miniaturas SET image_url=?, name=?, arrival_date=?, stock=?, price=?, observations=?, max_reservations_per_user=? WHERE id=?",
                  (image_url, name, arrival_date, stock, price, observations, max_reservations_per_user, mini_id))
        conn.commit()
        create_json_backup() # Backup local
        backup_to_google() # Backup Google Sheets
        sync_to_sheets() # Sincroniza planilha principal
        conn.close()
        flash('Miniatura atualizada com sucesso!', 'success')
        return redirect(url_for('admin'))
    
    conn.close()
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Editar Miniatura - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS Admin</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/admin" class="text-slate-300 hover:text-blue-400">Voltar para Admin</a>
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Editar Miniatura</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            <div class="bg-slate-800 rounded-lg p-6 max-w-lg mx-auto border-2 border-blue-600">
                <form method="POST" class="space-y-4">
                    <div>
                        <label for="image_url" class="block text-slate-300 font-bold mb-1">URL da Imagem:</label>
                        <input type="text" id="image_url" name="image_url" value="{{ miniatura[0] }}" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="name" class="block text-slate-300 font-bold mb-1">Nome da Miniatura:</label>
                        <input type="text" id="name" name="name" value="{{ miniatura[1] }}" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="arrival_date" class="block text-slate-300 font-bold mb-1">Previsão de Chegada:</label>
                        <input type="text" id="arrival_date" name="arrival_date" value="{{ miniatura[2] }}" class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="stock" class="block text-slate-300 font-bold mb-1">Estoque:</label>
                        <input type="number" id="stock" name="stock" value="{{ miniatura[3] }}" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="price" class="block text-slate-300 font-bold mb-1">Preço:</label>
                        <input type="number" step="0.01" id="price" name="price" value="{{ miniatura[4] }}" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="observations" class="block text-slate-300 font-bold mb-1">Observações:</label>
                        <textarea id="observations" name="observations" rows="3" class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">{{ miniatura[5] }}</textarea>
                    </div>
                    <div>
                        <label for="max_reservations_per_user" class="block text-slate-300 font-bold mb-1">Máx. Reservas por Usuário:</label>
                        <input type="number" id="max_reservations_per_user" name="max_reservations_per_user" value="{{ miniatura[6] }}" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <button type="submit" class="w-full bg-green-600 text-white font-bold py-2 rounded-lg hover:bg-green-700 transition duration-300">Atualizar Miniatura</button>
                </form>
            </div>
        </div>
    </body>
    </html>
    ''', miniatura=miniatura, LOGO_URL=LOGO_URL)

@app.route('/admin/delete-miniatura/<int:mini_id>')
@login_required
@admin_required
def delete_miniatura(mini_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM miniaturas WHERE id = ?", (mini_id,))
    conn.commit()
    create_json_backup() # Backup local
    backup_to_google() # Backup Google Sheets
    sync_to_sheets() # Sincroniza planilha principal
    conn.close()
    flash('Miniatura excluída com sucesso!', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/gerenciar-usuarios')
@login_required
@admin_required
def gerenciar_usuarios():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT id, name, email, phone, is_admin FROM users ORDER BY name')
    users = c.fetchall()
    conn.close()
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Gerenciar Usuários - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS Admin</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/admin" class="text-slate-300 hover:text-blue-400">Voltar para Admin</a>
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Gerenciar Usuários</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}

            {% if users %}
                <div class="overflow-x-auto bg-slate-800 rounded-lg border-2 border-blue-600">
                    <table class="min-w-full divide-y divide-slate-700">
                        <thead class="bg-slate-700">
                            <tr>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">ID</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Nome</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Email</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Telefone</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Admin</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Ações</th>
                            </tr>
                        </thead>
                        <tbody class="bg-slate-800 divide-y divide-slate-700">
                            {% for user in users %}
                                <tr>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-white">{{ user[0] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ user[1] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ user[2] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ user[3] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ 'Sim' if user[4] else 'Não' }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                        {% if user[2] != 'admin@jgminis.com.br' %} {# Não permite editar o admin principal #}
                                            <a href="/admin/toggle-admin/{{ user[0] }}" class="text-indigo-400 hover:text-indigo-600 mr-4">
                                                {{ 'Remover Admin' if user[4] else 'Tornar Admin' }}
                                            </a>
                                            <a href="/admin/delete-user/{{ user[0] }}" class="text-red-400 hover:text-red-600">Excluir</a>
                                        {% else %}
                                            <span class="text-slate-500">N/A</span>
                                        {% endif %}
                                    </td>
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            {% else %}
                <p class="text-white text-center text-lg">Nenhum usuário cadastrado.</p>
            {% endif %}
        </div>
    </body>
    </html>
    ''', users=users, LOGO_URL=LOGO_URL)

@app.route('/admin/toggle-admin/<int:user_id>')
@login_required
@admin_required
def toggle_admin(user_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT email, is_admin FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()

    if not user:
        conn.close()
        flash('Usuário não encontrado.', 'error')
        return redirect(url_for('gerenciar_usuarios'))
    
    if user[0] == 'admin@jgminis.com.br':
        conn.close()
        flash('Não é possível alterar as permissões do administrador principal.', 'error')
        return redirect(url_for('gerenciar_usuarios'))

    new_admin_status = 1 if not user[1] else 0
    c.execute("UPDATE users SET is_admin = ? WHERE id = ?", (new_admin_status, user_id))
    conn.commit()
    create_json_backup() # Backup local
    conn.close()
    flash(f'Permissões de admin para {user[0]} alteradas para {"Sim" if new_admin_status else "Não"}.', 'success')
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/admin/delete-user/<int:user_id>')
@login_required
@admin_required
def delete_user(user_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT email FROM users WHERE id = ?", (user_id,))
    user_email = c.fetchone()[0]

    if user_email == 'admin@jgminis.com.br':
        conn.close()
        flash('Não é possível excluir o administrador principal.', 'error')
        return redirect(url_for('gerenciar_usuarios'))

    c.execute("DELETE FROM users WHERE id = ?", (user_id,))
    c.execute("DELETE FROM reservations WHERE user_id = ?", (user_id,)) # Exclui reservas do usuário
    c.execute("DELETE FROM waitlist WHERE user_id = ?", (user_id,)) # Exclui lista de espera do usuário
    conn.commit()
    create_json_backup() # Backup local
    conn.close()
    flash(f'Usuário {user_email} e suas reservas/listas de espera excluídos com sucesso!', 'success')
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/admin/export-reservas')
@login_required
@admin_required
def export_reservas():
    buffer = export_to_excel_reservas()
    if buffer:
        return send_file(buffer, as_attachment=True, download_name='reservas_jgminis.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    flash('Nenhuma reserva para exportar.', 'info')
    return redirect(url_for('admin'))

@app.route('/admin/export-usuarios')
@login_required
@admin_required
def export_usuarios():
    buffer = export_to_excel_usuarios()
    if buffer:
        return send_file(buffer, as_attachment=True, download_name='usuarios_jgminis.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    flash('Nenhum usuário para exportar.', 'info')
    return redirect(url_for('admin'))

@app.route('/admin/send-broadcast', methods=['GET', 'POST'])
@login_required
@admin_required
def send_broadcast():
    if request.method == 'POST':
        subject = request.form['subject']
        message_body = request.form['message_body']
        
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT email FROM users")
        users = c.fetchall()
        conn.close()
        
        sent_count = 0
        for user in users:
            corpo_html = f'''
            <h1>{subject}</h1>
            <p>Olá,</p>
            <p>{message_body}</p>
            <p>Atenciosamente,<br>Equipe JG MINIS</p>
            '''
            if enviar_email(user[0], subject, corpo_html):
                sent_count += 1
        
        flash(f'E-mail enviado para {sent_count} usuários!', 'success')
        return redirect(url_for('admin'))
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Enviar Broadcast - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS Admin</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/admin" class="text-slate-300 hover:text-blue-400">Voltar para Admin</a>
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Enviar E-mail Broadcast</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}
            <div class="bg-slate-800 rounded-lg p-6 max-w-lg mx-auto border-2 border-blue-600">
                <form method="POST" class="space-y-4">
                    <div>
                        <label for="subject" class="block text-slate-300 font-bold mb-1">Assunto:</label>
                        <input type="text" id="subject" name="subject" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500">
                    </div>
                    <div>
                        <label for="message_body" class="block text-slate-300 font-bold mb-1">Corpo da Mensagem (HTML):</label>
                        <textarea id="message_body" name="message_body" rows="8" required class="w-full px-4 py-2 rounded-lg bg-slate-700 text-white border-2 border-blue-600 focus:outline-none focus:border-red-500"></textarea>
                    </div>
                    <button type="submit" class="w-full bg-green-600 text-white font-bold py-2 rounded-lg hover:bg-green-700 transition duration-300">Enviar E-mail</button>
                </form>
            </div>
        </div>
    </body>
    </html>
    ''', LOGO_URL=LOGO_URL)

@app.route('/admin/waitlist')
@login_required
@admin_required
def admin_waitlist():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''SELECT w.id, u.name, m.name, w.email, w.request_date
                 FROM waitlist w JOIN users u ON w.user_id = u.id JOIN miniaturas m ON w.miniatura_id = m.id
                 ORDER BY w.request_date DESC''')
    waitlist = c.fetchall()
    conn.close()
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Lista de Espera - JG MINIS</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gradient-to-b from-slate-950 via-blue-950 to-black min-h-screen">
        <nav class="bg-slate-800 border-b-2 border-red-600 p-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center">
                    <img src="''' + LOGO_URL + '''" class="h-10 rounded-full border-2 border-blue-400 mr-4">
                    <h1 class="text-2xl font-black text-blue-400">JG MINIS Admin</h1>
                </div>
                <div class="flex space-x-4">
                    <a href="/admin" class="text-slate-300 hover:text-blue-400">Voltar para Admin</a>
                    <a href="/logout" class="text-red-400 hover:text-red-300">Logout</a>
                </div>
            </div>
        </nav>
        <div class="p-6">
            <h2 class="text-3xl font-black text-white mb-6">Gerenciar Lista de Espera</h2>
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <ul class="mb-4">
                        {% for category, message in messages %}
                            <li class="text-{{ 'red' if category == 'error' else 'green' }}-400 text-center">{{ message }}</li>
                        {% endfor %}
                    </ul>
                {% endif %}
            {% endwith %}

            {% if waitlist %}
                <div class="overflow-x-auto bg-slate-800 rounded-lg border-2 border-blue-600">
                    <table class="min-w-full divide-y divide-slate-700">
                        <thead class="bg-slate-700">
                            <tr>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">ID</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Usuário</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Miniatura</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Email</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Data Pedido</th>
                                <th scope="col" class="px-6 py-3 text-left text-xs font-medium text-slate-300 uppercase tracking-wider">Ações</th>
                            </tr>
                        </thead>
                        <tbody class="bg-slate-800 divide-y divide-slate-700">
                            {% for item in waitlist %}
                                <tr>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-white">{{ item[0] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ item[1] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ item[2] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ item[3] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-sm text-slate-300">{{ item[4] }}</td>
                                    <td class="px-6 py-4 whitespace-nowrap text-right text-sm font-medium">
                                        <a href="/admin/clear-waitlist/{{ item[0] }}" class="text-red-400 hover:text-red-600">Remover</a>
                                    </td>
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            {% else %}
                <p class="text-white text-center text-lg">Ninguém na lista de espera.</p>
            {% endif %}
        </div>
    </body>
    </html>
    ''', waitlist=waitlist, LOGO_URL=LOGO_URL)

@app.route('/admin/clear-waitlist/<int:waitlist_id>')
@login_required
@admin_required
def clear_waitlist(waitlist_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM waitlist WHERE id = ?", (waitlist_id,))
    conn.commit()
    create_json_backup() # Backup local
    conn.close()
    flash('Item da lista de espera removido com sucesso!', 'success')
    return redirect(url_for('admin_waitlist'))

@app.route('/admin/load-sheets')
@login_required
@admin_required
def load_sheets_admin():
    load_miniaturas_from_sheets()
    flash('Miniaturas carregadas/atualizadas do Google Sheets!', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/manual-backup')
@login_required
@admin_required
def manual_backup_admin():
    create_json_backup()
    backup_to_google()
    sync_to_sheets()
    flash('Backup manual (JSON local + Google Sheets) e sincronização com planilha principal realizados!', 'success')
    return redirect(url_for('admin'))

# Inicialização leve (sem Sheets no boot)
init_db()
restore_backup()
load_initial_data()
# load_miniaturas_from_sheets() # Comente – mova para runtime
# create_json_backup() # Comente – mova para runtime

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=False)
